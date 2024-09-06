import markdown_pdf
import openpyxl

from domain.cocktail import Cocktail, CocktailIngredient

FILE_NAME = 'cocktails.xlsx'
COCKTAILS_TABLE_NAME = 'Receitas'
INGREDIENTS_TABLE_NAME = 'Ingredientes'
MENU_TITLE = 'Carta de Coquetéis'


def read_cocktails_from_excel(file_path: str, table_name: str) -> list[Cocktail]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[table_name]
    cocktails: list[Cocktail] = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        preparation = row[0]
        ice_type = row[1] 
        garnish = row[2]
        glassware = row[3]
        page = row[4]
        category = row[5]
        observation = row[6]
        name = row[7]
        
        ingredients = []
        
        if name is None:
            continue
        
        category = category.strip()
        name = name.strip()
        
        for i in range(8, len(row), 2):
            if row[i] is None:
                break
            
            ingredient_name = row[i].strip()
            ingredient_quantity = row[i + 1]
            ingredients.append(CocktailIngredient(ingredient_name, ingredient_quantity))
        
        cocktails.append(Cocktail(preparation, garnish, glassware, name, category, ingredients, page, ice_type, observation))
        
    return cocktails

def generate_menu(cocktails: list[Cocktail]) -> markdown_pdf.MarkdownPdf:
    categories: dict[str, list[Cocktail]] = {}
    pdf = markdown_pdf.MarkdownPdf(toc_level=6)
    
    for cocktail in cocktails:
        if cocktail.category not in categories:
            categories[cocktail.category] = []
        
        categories[cocktail.category].append(cocktail)
    
    index_text = f'# {MENU_TITLE}\n'
    for category in categories.keys():
        index_text += f"## {category}\n"

        for cocktail in categories[category]:
            max_page_chars = 3
            page_str = f'{cocktail.page}'.rjust(max_page_chars, '0')
            index_text += f'- {page_str} {cocktail.name}\n'

    pdf.add_section(
        markdown_pdf.Section(index_text), user_css="body { font-family: Inconsolata, monospace; }"
    )
        
    for category, cocktails in categories.items():
        markdown_text = f"## {category}\n"
        
        for cocktail in cocktails:            
            markdown_text += f'### {cocktail.name}\n'
            
            for ingredient in cocktail.ingredients:
                markdown_text += f"- {ingredient.name}: {ingredient.quantity}\n"
            
            markdown_text += "\n"  
            
            markdown_text += f"**Preparo:** {cocktail.preparation}\\\n"
            markdown_text += f"**Guarnição:** {cocktail.garnish}\\\n"
            markdown_text += f"**Copo:** {cocktail.glassware}\\\n"
            markdown_text += f"**Página:** {cocktail.page}"
            
            if cocktail.ice_type is not None or cocktail.observation is not None:
                markdown_text += "\\\n"
            else:
                markdown_text += "\n"
            
            if cocktail.ice_type is not None:
                markdown_text += f"**Tipo de gelo:** {cocktail.ice_type}"
                
                if cocktail.observation is not None:
                    markdown_text += "\\\n"
                else:
                    markdown_text += "\n"
                
            if cocktail.observation is not None:
                markdown_text += f"**Observação**: {cocktail.observation}\n"
        
        pdf.add_section(markdown_pdf.Section(markdown_text), user_css="body { font-family: Inconsolata, monospace; }")
    
    return pdf


def read_availible_ingridients_from_excel(file_path: str, table_name: str) -> list[str]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[table_name]
    ingredients: list[str] = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        ingredient = row[0]
        if ingredient is None:
            continue
        
        ingredient = ingredient.strip()
        
        is_availible = row[1]
        if is_availible is not None:
            ingredients.append(ingredient)
            
    return ingredients

def get_availible_cocktails(file_path: str, cocktails_table: str, ingredients_table: str) -> list[Cocktail]:
    cocktails = read_cocktails_from_excel(file_path, cocktails_table)
    availible_ingredients = read_availible_ingridients_from_excel(file_path, ingredients_table)
    
    availible_cocktail = []
    for cocktail in cocktails:
        is_availible = True
        for ingredient in cocktail.ingredients:
            if ingredient.name not in availible_ingredients:
                is_availible = False
                break
        
        if is_availible:
            availible_cocktail.append(cocktail)
            
    return availible_cocktail


def main():
    cocktails = get_availible_cocktails(FILE_NAME, COCKTAILS_TABLE_NAME, INGREDIENTS_TABLE_NAME)
    pdf = generate_menu(cocktails)
    pdf.save("menu.pdf")
        
if __name__ == '__main__':
    main()
    